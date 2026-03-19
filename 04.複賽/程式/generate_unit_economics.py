import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === Styles ===
font_title = Font(name='Arial', size=16, bold=True, color='1F4E79')
font_section = Font(name='Arial', size=13, bold=True, color='2E75B6')
font_subsection = Font(name='Arial', size=11, bold=True, color='404040')
font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
font_input = Font(name='Arial', size=10, color='0000FF')  # Blue = user input
font_calc = Font(name='Arial', size=10, color='000000')   # Black = formula
font_link = Font(name='Arial', size=10, color='008000')   # Green = cross-sheet link
font_label = Font(name='Arial', size=10, color='404040')
font_note = Font(name='Arial', size=9, italic=True, color='808080')
font_bold = Font(name='Arial', size=10, bold=True)
font_result = Font(name='Arial', size=11, bold=True, color='1F4E79')
font_check = Font(name='Arial', size=10, bold=True, color='008000')
font_warn = Font(name='Arial', size=10, bold=True, color='FF0000')

fill_input = PatternFill('solid', fgColor='FFF2CC')      # Light yellow = input cell
fill_header = PatternFill('solid', fgColor='2E75B6')      # Blue header
fill_section = PatternFill('solid', fgColor='D6E4F0')     # Light blue section
fill_result = PatternFill('solid', fgColor='E2EFDA')      # Light green result
fill_light = PatternFill('solid', fgColor='F2F2F2')       # Light gray alternating
fill_white = PatternFill('solid', fgColor='FFFFFF')
fill_scenario_good = PatternFill('solid', fgColor='C6EFCE')
fill_scenario_base = PatternFill('solid', fgColor='FFEB9C')
fill_scenario_bad = PatternFill('solid', fgColor='FFC7CE')

align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
bottom_border = Border(bottom=Side(style='medium', color='2E75B6'))

NT_FMT = '#,##0'
NT_FULL = '"NT$"#,##0'
PCT_FMT = '0.0%'
RATIO_FMT = '0.0'
MON_FMT = '0.0'

def style_input_cell(ws, row, col, value=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_input
    c.fill = fill_input
    c.border = thin_border
    c.alignment = align_right
    if fmt: c.number_format = fmt
    return c

def style_calc_cell(ws, row, col, value=None, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_bold if bold else font_calc
    c.border = thin_border
    c.alignment = align_right
    if fmt: c.number_format = fmt
    return c

def style_label(ws, row, col, value, indent=0):
    c = ws.cell(row=row, column=col, value=('  '*indent + value if indent else value))
    c.font = font_label
    c.border = thin_border
    c.alignment = align_left
    return c

def style_header_row(ws, row, cols, values):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=cols[0]+i, value=v)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = thin_border

def style_section_row(ws, row, col_start, col_end, value):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.font = font_section
    c.fill = fill_section
    c.alignment = align_left
    for col in range(col_start, col_end+1):
        ws.cell(row=row, column=col).fill = fill_section
        ws.cell(row=row, column=col).border = thin_border

def style_result_row(ws, row, col, label, val_col, fmt=None):
    c1 = ws.cell(row=row, column=col, value=label)
    c1.font = font_result
    c1.fill = fill_result
    c1.border = thin_border
    c2 = ws.cell(row=row, column=val_col)
    c2.font = font_result
    c2.fill = fill_result
    c2.border = thin_border
    if fmt: c2.number_format = fmt
    for cc in range(col, val_col+1):
        ws.cell(row=row, column=cc).fill = fill_result
        ws.cell(row=row, column=cc).border = thin_border

# =========================================================
# SHEET 1: 假設參數 (Assumptions)
# =========================================================
ws1 = wb.active
ws1.title = '假設參數'
ws1.sheet_properties.tabColor = '2E75B6'

ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 40

r = 1
ws1.merge_cells('B1:E1')
c = ws1.cell(row=1, column=2, value='Mathify Labs — 假設參數儀表板')
c.font = font_title

r = 3
style_section_row(ws1, r, 2, 5, '定價與營收假設')
r = 4
style_header_row(ws1, r, [2,3,4,5], ['參數', '數值', '單位', '說明'])
params = [
    (5, '月費', 600, 'NT$/月', '付費用戶每月訂閱費'),
    (6, '年繳折扣', 0.83, '', '年繳方案為月費 × 10（83折）'),
    (7, '年繳用戶占比', 0.0, '', '目前假設 0%，可調整'),
]
for row, label, val, unit, note in params:
    style_label(ws1, row, 2, label)
    style_input_cell(ws1, row, 3, val, NT_FMT if isinstance(val, int) else PCT_FMT if val < 1 else NT_FMT)
    ws1.cell(row=row, column=4, value=unit).font = font_label
    ws1.cell(row=row, column=5, value=note).font = font_note
    ws1.cell(row=row, column=5).alignment = align_wrap

r = 8
style_label(ws1, r, 2, 'ARPPU（付費用戶月均）')
style_calc_cell(ws1, r, 3, '=C5*(1-C7*(1-C6))', NT_FMT)
ws1.cell(row=r, column=4, value='NT$/月').font = font_label
ws1.cell(row=r, column=5, value='考慮年繳折扣後的付費用戶實際月均收入').font = font_note

r = 9
style_label(ws1, r, 2, '免費→付費轉換率')
style_input_cell(ws1, r, 3, 0.033, '0.0%')
ws1.cell(row=r, column=5, value='Year 1 目標：500 付費 / 15,000 總用戶 ≈ 3.3%').font = font_note
ws1.cell(row=r, column=5).alignment = align_wrap

r = 11
style_section_row(ws1, r, 2, 5, '每用戶變動成本（COGS）')
r = 12
style_header_row(ws1, r, [2,3,4,5], ['成本項目', '數值', '單位', '說明'])
costs = [
    (13, 'AI 推理成本', 100, 'NT$/月/用戶', '平均 150 次查詢/月 × NT$0.6/次'),
    (14, '雲端基礎設施', 20, 'NT$/月/用戶', '伺服器、CDN、資料庫分攤'),
]
for row, label, val, unit, note in costs:
    style_label(ws1, row, 2, label)
    style_input_cell(ws1, row, 3, val, NT_FMT)
    ws1.cell(row=row, column=4, value=unit).font = font_label
    ws1.cell(row=row, column=5, value=note).font = font_note
    ws1.cell(row=row, column=5).alignment = align_wrap

r = 15
style_label(ws1, r, 2, 'COGS 合計')
style_calc_cell(ws1, r, 3, '=C13+C14', NT_FMT, bold=True)
ws1.cell(row=r, column=4, value='NT$/月/用戶').font = font_label

r = 17
style_section_row(ws1, r, 2, 5, '客戶生命週期假設')
r = 18
style_header_row(ws1, r, [2,3,4,5], ['學生類型', '占比', '留存月數', '說明'])
cohorts = [
    (19, '短期使用（1 學年）', 0.60, 10, '段考/會考衝刺後離開'),
    (20, '中期使用（2 學年）', 0.30, 20, '跨學年持續訂閱'),
    (21, '長期使用（3 年+）', 0.10, 30, '國中→高中續用'),
]
for row, label, pct, months, note in cohorts:
    style_label(ws1, row, 2, label)
    style_input_cell(ws1, row, 3, pct, PCT_FMT)
    style_input_cell(ws1, row, 4, months, '0')
    ws1.cell(row=row, column=5, value=note).font = font_note
    ws1.cell(row=row, column=5).alignment = align_wrap

r = 22
style_label(ws1, r, 2, '加權平均生命週期')
style_calc_cell(ws1, r, 3, '=SUMPRODUCT(C19:C21,D19:D21)', MON_FMT, bold=True)
ws1.cell(row=r, column=4, value='個月').font = font_label

r = 23
style_label(ws1, r, 2, '保守調整係數')
style_input_cell(ws1, r, 3, 0.93, PCT_FMT)
ws1.cell(row=r, column=5, value='考量寒暑假流失，15→14 個月 ≈ 93%').font = font_note

r = 24
style_label(ws1, r, 2, '調整後生命週期')
style_calc_cell(ws1, r, 3, '=ROUND(C22*C23,0)', '0', bold=True)
ws1.cell(row=r, column=4, value='個月').font = font_label

r = 26
style_section_row(ws1, r, 2, 5, 'CAC（獲客成本）假設')
r = 27
style_header_row(ws1, r, [2,3,4,5], ['獲客渠道', 'CAC', '占比', '說明'])
channels = [
    (28, '免費轉付費（PLG）', 100, 0.40, '免費用戶自然轉換'),
    (29, '口碑推薦', 400, 0.25, '推薦獎勵機制'),
    (30, 'KOL / 社群行銷', 1000, 0.20, '數學補教 YouTuber 合作'),
    (31, '校園推廣', 650, 0.15, '教育展、學校免費試用'),
]
for row, label, cac, pct, note in channels:
    style_label(ws1, row, 2, label)
    style_input_cell(ws1, row, 3, cac, NT_FMT)
    style_input_cell(ws1, row, 4, pct, PCT_FMT)
    ws1.cell(row=row, column=5, value=note).font = font_note

r = 32
style_label(ws1, r, 2, '加權平均 CAC')
style_calc_cell(ws1, r, 3, '=SUMPRODUCT(C28:C31,D28:D31)', NT_FMT, bold=True)
ws1.cell(row=r, column=4, value='NT$').font = font_label

r = 33
style_label(ws1, r, 2, 'CAC 緩衝（含測試成本）')
style_input_cell(ws1, r, 3, 60, NT_FMT)
ws1.cell(row=r, column=5, value='渠道測試與優化額外支出').font = font_note

r = 34
style_label(ws1, r, 2, '最終 CAC')
style_calc_cell(ws1, r, 3, '=C32+C33', NT_FMT, bold=True)

r = 36
style_section_row(ws1, r, 2, 5, '用戶成長假設')
r = 37
style_header_row(ws1, r, [2,3,4,5], ['年度', '付費用戶數', '免費用戶數（年底）', '付費啟動月'])
years_data = [
    (38, 'Year 1', 500, 15000, 6),
    (39, 'Year 2', 2000, 40000, 1),
    (40, 'Year 3', 5000, 80000, 1),
]
for row, label, paid, free, start_m in years_data:
    style_label(ws1, row, 2, label)
    style_input_cell(ws1, row, 3, paid, '#,##0')
    style_input_cell(ws1, row, 4, free, '#,##0')
    style_input_cell(ws1, row, 5, start_m, '0')

# Conversion rate (calculated)
r = 41
style_label(ws1, r, 2, '實際轉換率（驗算）')
style_calc_cell(ws1, r, 3, '=IF(C38+D38=0,0,C38/(C38+D38))', '0.0%')
style_calc_cell(ws1, r, 4, '=IF(C39+D39=0,0,C39/(C39+D39))', '0.0%')
style_calc_cell(ws1, r, 5, '=IF(C40+D40=0,0,C40/(C40+D40))', '0.0%')

# ARPU (true, all users)
r = 42
style_label(ws1, r, 2, 'ARPU（全用戶月均）')
style_calc_cell(ws1, r, 3, '=C8*C41', NT_FMT)
style_calc_cell(ws1, r, 4, '=C8*D41', NT_FMT)
style_calc_cell(ws1, r, 5, '=C8*E41', NT_FMT)
ws1.cell(row=r, column=3).font = font_calc

r = 44
style_section_row(ws1, r, 2, 5, '營運成本假設')
r = 45
style_header_row(ws1, r, [2,3,4,5], ['項目', 'Year 1', 'Year 2', 'Year 3'])
opex_items = [
    (46, '月固定成本', 160000, 350000, 700000, '人事+辦公+行政（不含 COGS）'),
    (47, '創業支持金', 3000000, 0, 0, 'NT$300 萬一次性撥款'),
]
for row, label, y1, y2, y3, note in opex_items:
    style_label(ws1, row, 2, label)
    style_input_cell(ws1, row, 3, y1, NT_FMT)
    style_input_cell(ws1, row, 4, y2, NT_FMT)
    style_input_cell(ws1, row, 5, y3 if y3 else None, NT_FMT)
    if row == 46:
        ws1.cell(row=row, column=5, value=y3)
        ws1.cell(row=row, column=5).font = font_input
        ws1.cell(row=row, column=5).fill = fill_input
        ws1.cell(row=row, column=5).number_format = NT_FMT

# =========================================================
# SHEET 2: 單位經濟 (Unit Economics)
# =========================================================
ws2 = wb.create_sheet('單位經濟')
ws2.sheet_properties.tabColor = '008000'

ws2.column_dimensions['A'].width = 4
ws2.column_dimensions['B'].width = 24
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 14

r = 1
ws2.merge_cells('B1:E1')
ws2.cell(row=1, column=2, value='單位經濟模型（自動計算）').font = font_title

r = 3
style_section_row(ws2, r, 2, 5, 'ARPPU vs ARPU')
r = 4
style_label(ws2, r, 2, 'ARPPU（付費用戶月均）')
style_calc_cell(ws2, r, 3, "=假設參數!C8", NT_FULL)
ws2.cell(row=r, column=3).font = font_link

r = 5
style_label(ws2, r, 2, '免費→付費轉換率')
style_calc_cell(ws2, r, 3, "=假設參數!C9", PCT_FMT)
ws2.cell(row=r, column=3).font = font_link

r = 6
style_label(ws2, r, 2, 'ARPU（全用戶月均）')
c = style_calc_cell(ws2, r, 3, '=C4*C5', NT_FULL)
c.font = font_result
ws2.cell(row=r, column=4, value='= ARPPU × 轉換率').font = font_note

r = 8
style_section_row(ws2, r, 2, 5, '付費用戶成本結構')
r = 9
style_label(ws2, r, 2, 'COGS（月/付費用戶）')
style_calc_cell(ws2, r, 3, "=假設參數!C15", NT_FULL)
ws2.cell(row=r, column=3).font = font_link

r = 10
style_label(ws2, r, 2, '單位貢獻利潤')
c = style_calc_cell(ws2, r, 3, '=C4-C9', NT_FULL)
c.font = font_result

r = 11
style_label(ws2, r, 2, '貢獻利潤率')
c = style_calc_cell(ws2, r, 3, '=IF(C4=0,0,C10/C4)', PCT_FMT)
c.font = font_result
ws2.cell(row=r, column=4, value='=IF(C11>=0.7,"✅ 健康","⚠️ 偏低")').font = font_check

r = 13
style_section_row(ws2, r, 2, 5, '客戶生命週期')
r = 14
style_label(ws2, r, 2, '加權平均生命週期')
style_calc_cell(ws2, r, 3, "=假設參數!C22", MON_FMT)
ws2.cell(row=r, column=3).font = font_link
ws2.cell(row=r, column=4, value='個月（原始）').font = font_note

r = 15
style_label(ws2, r, 2, '調整後生命週期')
style_calc_cell(ws2, r, 3, "=假設參數!C24", '0')
ws2.cell(row=r, column=3).font = font_link
ws2.cell(row=r, column=4, value='個月（保守值）').font = font_note

r = 17
style_section_row(ws2, r, 2, 5, 'LTV（客戶終身價值）')
r = 18
style_label(ws2, r, 2, 'LTV')
c = style_calc_cell(ws2, r, 3, '=C10*C15', NT_FULL)
c.font = font_result
c.fill = fill_result
for cc in [2,3,4]:
    ws2.cell(row=r, column=cc).fill = fill_result
ws2.cell(row=r, column=4, value='= 貢獻利潤 × 生命週期').font = font_note

r = 20
style_section_row(ws2, r, 2, 5, 'CAC（獲客成本）')
r = 21
style_label(ws2, r, 2, 'CAC')
style_calc_cell(ws2, r, 3, "=假設參數!C34", NT_FULL)
ws2.cell(row=r, column=3).font = font_link

r = 23
style_header_row(ws2, r, [2,3,4,5], ['指標', '數值', '業界基準', '判斷'])
metrics = [
    (24, 'LTV / CAC', '=IF(C21=0,0,C18/C21)', RATIO_FMT, '>3:1', '=IF(C24>=3,"✅ 健康","❌ 不足")'),
    (25, 'CAC 回收期（月）', '=IF(C10=0,0,C21/C10)', MON_FMT, '<12 個月', '=IF(C25<=12,"✅ 健康","⚠️ 偏長")'),
    (26, '貢獻利潤率', '=C11', PCT_FMT, '>70%', '=IF(C26>=0.7,"✅ SaaS 健康","⚠️ 偏低")'),
]
for row, label, formula, fmt, benchmark, check_formula in metrics:
    style_label(ws2, row, 2, label)
    c = style_calc_cell(ws2, row, 3, formula, fmt)
    c.font = font_result
    ws2.cell(row=row, column=4, value=benchmark).font = font_label
    ws2.cell(row=row, column=4).alignment = align_center
    ws2.cell(row=row, column=5, value=check_formula).font = font_check

r = 28
style_section_row(ws2, r, 2, 5, '敏感度分析')
r = 29
style_header_row(ws2, r, [2,3,4,5], ['情境', '生命週期', 'CAC', 'LTV/CAC'])

# Optimistic
r = 30
ws2.cell(row=r, column=2, value='🟢 樂觀').font = font_bold
style_input_cell(ws2, r, 3, 18, '0')
style_input_cell(ws2, r, 4, 400, NT_FMT)
ws2.cell(row=r, column=5, value='=IF(D30=0,0,(C4-C9)*C30/D30)').font = font_calc
ws2.cell(row=r, column=5).number_format = RATIO_FMT
for cc in [2,3,4,5]: ws2.cell(row=r, column=cc).fill = fill_scenario_good

# Baseline
r = 31
ws2.cell(row=r, column=2, value='🟡 基準').font = font_bold
style_calc_cell(ws2, r, 3, '=C15', '0')
style_calc_cell(ws2, r, 4, '=C21', NT_FMT)
ws2.cell(row=r, column=5, value='=IF(D31=0,0,(C4-C9)*C31/D31)').font = font_calc
ws2.cell(row=r, column=5).number_format = RATIO_FMT
for cc in [2,3,4,5]: ws2.cell(row=r, column=cc).fill = fill_scenario_base

# Pessimistic
r = 32
ws2.cell(row=r, column=2, value='🔴 悲觀').font = font_bold
style_input_cell(ws2, r, 3, 10, '0')
style_input_cell(ws2, r, 4, 700, NT_FMT)
ws2.cell(row=r, column=5, value='=IF(D32=0,0,(C4-C9)*C32/D32)').font = font_calc
ws2.cell(row=r, column=5).number_format = RATIO_FMT
for cc in [2,3,4,5]: ws2.cell(row=r, column=cc).fill = fill_scenario_bad

# =========================================================
# SHEET 3: 財務預測 (Financial Projections)
# =========================================================
ws3 = wb.create_sheet('三年財務預測')
ws3.sheet_properties.tabColor = 'FF6600'

ws3.column_dimensions['A'].width = 4
ws3.column_dimensions['B'].width = 26
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 18

r = 1
ws3.merge_cells('B1:E1')
ws3.cell(row=1, column=2, value='三年財務預測（自動連動）').font = font_title

# --- Revenue Section ---
r = 3
style_section_row(ws3, r, 2, 5, '營收預測')
r = 4
style_header_row(ws3, r, [2,3,4,5], ['項目', 'Year 1', 'Year 2', 'Year 3'])

r = 5
style_label(ws3, r, 2, '付費用戶數（年底）')
for i, col in enumerate([3,4,5]):
    style_calc_cell(ws3, r, col, f"=假設參數!C{38+i}", '#,##0')
    ws3.cell(row=r, column=col).font = font_link

r = 6
style_label(ws3, r, 2, '免費用戶數（年底）')
for i, col in enumerate([3,4,5]):
    style_calc_cell(ws3, r, col, f"=假設參數!D{38+i}", '#,##0')
    ws3.cell(row=r, column=col).font = font_link

r = 7
style_label(ws3, r, 2, '全部用戶數（年底）')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'={cl}5+{cl}6', '#,##0')

r = 8
style_label(ws3, r, 2, '轉換率')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'=IF({cl}7=0,0,{cl}5/{cl}7)', '0.0%')

r = 9
style_label(ws3, r, 2, 'ARPPU（月）')
for col in [3,4,5]:
    style_calc_cell(ws3, r, col, "=假設參數!C8", NT_FULL)
    ws3.cell(row=r, column=col).font = font_link

r = 10
style_label(ws3, r, 2, 'ARPU（全用戶月均）')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'={cl}9*{cl}8', NT_FULL)

r = 11
style_label(ws3, r, 2, '月均付費用戶')
ws3.cell(row=r, column=3, value='=ROUND(C5/2*(13-假設參數!E38)/12,0)').font = font_calc
ws3.cell(row=r, column=3).number_format = '#,##0'
ws3.cell(row=r, column=3).border = thin_border
ws3.cell(row=r, column=4, value='=ROUND((C5+D5)/2,0)').font = font_calc
ws3.cell(row=r, column=4).number_format = '#,##0'
ws3.cell(row=r, column=4).border = thin_border
ws3.cell(row=r, column=5, value='=ROUND((D5+E5)/2,0)').font = font_calc
ws3.cell(row=r, column=5).number_format = '#,##0'
ws3.cell(row=r, column=5).border = thin_border

r = 12
style_label(ws3, r, 2, '年度營收')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'={cl}11*{cl}9*12', '#,##0')
    ws3.cell(row=r, column=col).font = font_result

r = 13
style_label(ws3, r, 2, '營收成長率')
ws3.cell(row=r, column=3, value='—').font = font_calc
ws3.cell(row=r, column=3).alignment = align_right
ws3.cell(row=r, column=3).border = thin_border
ws3.cell(row=r, column=4, value='=IF(C12=0,0,(D12-C12)/C12)').font = font_calc
ws3.cell(row=r, column=4).number_format = '+0%'
ws3.cell(row=r, column=4).border = thin_border
ws3.cell(row=r, column=5, value='=IF(D12=0,0,(E12-D12)/D12)').font = font_calc
ws3.cell(row=r, column=5).number_format = '+0%'
ws3.cell(row=r, column=5).border = thin_border

# --- Cost Section ---
r = 15
style_section_row(ws3, r, 2, 5, '成本結構')
r = 16
style_header_row(ws3, r, [2,3,4,5], ['項目', 'Year 1', 'Year 2', 'Year 3'])

r = 17
style_label(ws3, r, 2, '變動成本（COGS）')
for i, col in enumerate([3,4,5]):
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'={cl}11*假設參數!C15*12', '#,##0')

r = 18
style_label(ws3, r, 2, '固定成本（OpEx）')
for i, col in enumerate([3,4,5]):
    style_calc_cell(ws3, r, col, f'=假設參數!{get_column_letter(3+i)}46*12', '#,##0')
    ws3.cell(row=r, column=col).font = font_link

r = 19
style_label(ws3, r, 2, '總成本')
for col in [3,4,5]:
    cl = get_column_letter(col)
    c = style_calc_cell(ws3, r, col, f'={cl}17+{cl}18', '#,##0')
    c.font = font_bold

# --- P&L ---
r = 21
style_section_row(ws3, r, 2, 5, '損益')
r = 22
style_label(ws3, r, 2, '毛利（營收 - COGS）')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'={cl}12-{cl}17', '#,##0')

r = 23
style_label(ws3, r, 2, '毛利率')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'=IF({cl}12=0,0,{cl}22/{cl}12)', PCT_FMT)

r = 24
style_label(ws3, r, 2, '淨利/虧損')
for col in [3,4,5]:
    cl = get_column_letter(col)
    c = style_calc_cell(ws3, r, col, f'={cl}12-{cl}19', '#,##0;(#,##0);"-"')
    c.font = font_result

r = 25
style_label(ws3, r, 2, '淨利率')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'=IF({cl}12=0,0,{cl}24/{cl}12)', '+0%;-0%')

# --- Cash Flow ---
r = 27
style_section_row(ws3, r, 2, 5, '現金流量')
r = 28
style_header_row(ws3, r, [2,3,4,5], ['項目', 'Year 1', 'Year 2', 'Year 3'])

r = 29
style_label(ws3, r, 2, '期初現金')
style_calc_cell(ws3, r, 3, '=假設參數!C47', '#,##0')
ws3.cell(row=r, column=3).font = font_link
style_calc_cell(ws3, r, 4, '=C31', '#,##0')
style_calc_cell(ws3, r, 5, '=D31', '#,##0')

r = 30
style_label(ws3, r, 2, '淨現金流')
for col in [3,4,5]:
    cl = get_column_letter(col)
    style_calc_cell(ws3, r, col, f'={cl}24', '#,##0;(#,##0)')

r = 31
style_label(ws3, r, 2, '期末現金')
for col in [3,4,5]:
    cl = get_column_letter(col)
    c = style_calc_cell(ws3, r, col, f'={cl}29+{cl}30', '#,##0;(#,##0)')
    c.font = font_result

# --- Break-even ---
r = 33
style_section_row(ws3, r, 2, 5, '損益兩平分析')
r = 34
style_label(ws3, r, 2, '月固定成本')
style_calc_cell(ws3, r, 3, '=假設參數!C46', NT_FULL)
ws3.cell(row=r, column=3).font = font_link

r = 35
style_label(ws3, r, 2, '每用戶月貢獻利潤')
style_calc_cell(ws3, r, 3, '=單位經濟!C10', NT_FULL)
ws3.cell(row=r, column=3).font = font_link

r = 36
style_label(ws3, r, 2, '損益兩平用戶數')
c = style_calc_cell(ws3, r, 3, '=IF(C35=0,0,ROUNDUP(C34/C35,0))', '#,##0')
c.font = font_result
c.fill = fill_result
ws3.cell(row=r, column=2).fill = fill_result
ws3.cell(row=r, column=4, value='名付費用戶').font = font_label

# --- MRR Milestones ---
r = 38
style_section_row(ws3, r, 2, 5, 'MRR 里程碑')
r = 39
style_header_row(ws3, r, [2,3,4,5], ['時間點', '付費用戶', 'MRR', '說明'])

r = 40
style_label(ws3, r, 2, 'M6')
style_input_cell(ws3, r, 3, 50, '#,##0')
style_calc_cell(ws3, r, 4, '=C40*假設參數!C8', NT_FULL)
ws3.cell(row=r, column=5, value='付費方案上線').font = font_note

r = 41
style_label(ws3, r, 2, 'M12（Year 1 結束）')
style_calc_cell(ws3, r, 3, '=假設參數!C38', '#,##0')
ws3.cell(row=r, column=3).font = font_link
style_calc_cell(ws3, r, 4, '=C41*假設參數!C8', NT_FULL)
ws3.cell(row=r, column=5, value='Year 1 目標').font = font_note

# =========================================================
# SHEET 4: 12 個月明細
# =========================================================
ws4 = wb.create_sheet('12個月明細')
ws4.sheet_properties.tabColor = '7030A0'

ws4.column_dimensions['A'].width = 4
ws4.column_dimensions['B'].width = 12

r = 1
ws4.merge_cells('B1:N1')
ws4.cell(row=1, column=2, value='Year 1 — 逐月營運預測').font = font_title

r = 3
headers = ['月份'] + [f'M{i}' for i in range(1, 13)]
for i, h in enumerate(headers):
    c = ws4.cell(row=r, column=2+i, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = thin_border
    if i > 0:
        ws4.column_dimensions[get_column_letter(2+i)].width = 13

# Row 4: Paying users (manual monthly targets matching milestones)
r = 4
style_label(ws4, r, 2, '付費用戶')
monthly_users = [0, 0, 0, 0, 0, 50, 125, 200, 250, 300, 400, 500]
for i, u in enumerate(monthly_users):
    style_input_cell(ws4, r, 3+i, u, '#,##0')

# Row 5: ARPU
r = 5
style_label(ws4, r, 2, 'ARPPU')
for i in range(12):
    style_calc_cell(ws4, r, 3+i, '=假設參數!C8', NT_FMT)  # ARPPU
    ws4.cell(row=r, column=3+i).font = font_link

# Row 6: Monthly revenue
r = 6
style_label(ws4, r, 2, '月營收')
for i in range(12):
    col = get_column_letter(3+i)
    style_calc_cell(ws4, r, 3+i, f'={col}4*{col}5', '#,##0')

# Row 7: COGS
r = 7
style_label(ws4, r, 2, 'COGS')
for i in range(12):
    col = get_column_letter(3+i)
    style_calc_cell(ws4, r, 3+i, f'={col}4*假設參數!C15', '#,##0')

# Row 8: Contribution
r = 8
style_label(ws4, r, 2, '貢獻利潤')
for i in range(12):
    col = get_column_letter(3+i)
    style_calc_cell(ws4, r, 3+i, f'={col}6-{col}7', '#,##0')

# Row 9: Fixed costs
r = 9
style_label(ws4, r, 2, '固定成本')
for i in range(12):
    style_calc_cell(ws4, r, 3+i, '=假設參數!C46', '#,##0')
    ws4.cell(row=r, column=3+i).font = font_link

# Row 10: Monthly P&L
r = 10
style_label(ws4, r, 2, '月損益')
for i in range(12):
    col = get_column_letter(3+i)
    c = style_calc_cell(ws4, r, 3+i, f'={col}8-{col}9', '#,##0;(#,##0)')
    c.font = font_bold

# Row 11: Cumulative P&L
r = 11
style_label(ws4, r, 2, '累計損益')
c = style_calc_cell(ws4, r, 3, '=C10', '#,##0;(#,##0)')
c.font = font_bold
for i in range(1, 12):
    col = get_column_letter(3+i)
    prev = get_column_letter(2+i)
    c = style_calc_cell(ws4, r, 3+i, f'={prev}11+{col}10', '#,##0;(#,##0)')
    c.font = font_bold

# Row 12: MRR (same as revenue for SaaS)
r = 12
style_label(ws4, r, 2, 'MRR')
for i in range(12):
    col = get_column_letter(3+i)
    c = style_calc_cell(ws4, r, 3+i, f'={col}6', NT_FULL)
    c.font = font_result

# Row 14: Summary
r = 14
ws4.merge_cells('B14:F14')
ws4.cell(row=14, column=2, value='Year 1 摘要').font = font_section

r = 15
style_label(ws4, r, 2, '全年營收合計')
style_calc_cell(ws4, r, 3, '=SUM(C6:N6)', '#,##0')
ws4.cell(row=r, column=3).font = font_result

r = 16
style_label(ws4, r, 2, 'M6 MRR')
style_calc_cell(ws4, r, 3, '=H6', NT_FULL)
ws4.cell(row=r, column=3).font = font_result

r = 17
style_label(ws4, r, 2, 'M12 MRR')
style_calc_cell(ws4, r, 3, '=N6', NT_FULL)
ws4.cell(row=r, column=3).font = font_result

# =========================================================
# Save
# =========================================================
output_path = r'C:\Users\226178\Desktop\GitHub\startup-asvda\04.複賽\Unit_Economics_Model.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
